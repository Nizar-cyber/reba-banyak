"""
REBA Ergonomics Analyzer — Streamlit Web App (Operator Daily)
Features:
- Login sederhana (session-based)
- Multi upload foto
- Input BEBAN & ACTIVITY SCORE per foto (editable table)
- Batch analisis + 1 file Excel hasil semua foto
- REBA lookup tables (Hignett & McAtamney, 2000)

Target environment:
- Streamlit Cloud
- Python 3.11
- mediapipe==0.10.21 (legacy mp.solutions tersedia)
"""

import cv2
import mediapipe as mp
import numpy as np
import pandas as pd
import streamlit as st
from datetime import datetime
from io import BytesIO
from PIL import Image

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="REBA Analyzer",
    page_icon="🦴",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
      footer {visibility: hidden;}
      #MainMenu {visibility: hidden;}
      .block-container {padding-top: 1.1rem; padding-bottom: 1.1rem;}
    </style>
    """,
    unsafe_allow_html=True,
)

# ─────────────────────────────────────────────────────────────────────────────
# LOGIN (simple session-based)
# ─────────────────────────────────────────────────────────────────────────────
LOGIN_USER = "CCY"
LOGIN_PASS = "Toyota2026"

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False


def _login_ui():
    st.title("🔐 Login REBA System")
    st.caption("Internal Kaizen Tool")
    u = st.text_input("Username")
    p = st.text_input("Password", type="password")
    if st.button("Login", type="primary", use_container_width=True):
        if u == LOGIN_USER and p == LOGIN_PASS:
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("❌ Username atau Password salah")


if not st.session_state.logged_in:
    _login_ui()
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# MEDIAPIPE (cache)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False)
def load_pose_model():
    mp_pose = mp.solutions.pose
    detector = mp_pose.Pose(static_image_mode=True, min_detection_confidence=0.5)
    return mp_pose, detector


mp_pose, pose_detector = load_pose_model()

# ─────────────────────────────────────────────────────────────────────────────
# REBA LOOKUP TABLES
# ─────────────────────────────────────────────────────────────────────────────
TABLE_A = [
    [[1, 2, 3, 4], [1, 2, 3, 4], [3, 3, 5, 6]],
    [[2, 3, 4, 5], [3, 4, 5, 6], [4, 5, 6, 7]],
    [[2, 4, 5, 6], [4, 5, 6, 7], [5, 6, 7, 8]],
    [[3, 5, 6, 7], [5, 6, 7, 8], [6, 7, 8, 9]],
    [[4, 6, 7, 8], [6, 7, 8, 9], [7, 8, 9, 9]],
]

TABLE_B = [
    [[1, 2, 2], [1, 2, 3]],
    [[1, 2, 3], [2, 3, 4]],
    [[3, 4, 5], [4, 5, 5]],
    [[4, 5, 5], [5, 6, 7]],
    [[6, 7, 8], [7, 8, 8]],
    [[7, 8, 8], [8, 9, 9]],
]

TABLE_C = [
    [1, 1, 1, 2, 3, 3, 4, 5, 6, 7, 7, 7],
    [1, 2, 2, 3, 4, 4, 5, 6, 6, 7, 7, 8],
    [2, 3, 3, 3, 4, 5, 6, 7, 7, 8, 8, 8],
    [3, 4, 4, 4, 5, 6, 7, 8, 8, 9, 9, 9],
    [4, 4, 4, 5, 6, 7, 8, 8, 9, 9, 10, 10],
    [6, 6, 6, 7, 8, 8, 9, 9, 10, 10, 11, 11],
    [7, 7, 7, 8, 9, 9, 10, 10, 11, 11, 11, 12],
    [8, 8, 8, 9, 10, 10, 11, 11, 12, 12, 13, 13],
    [9, 9, 9, 10, 10, 11, 11, 12, 13, 13, 13, 13],
    [10, 10, 10, 11, 11, 12, 12, 13, 13, 14, 14, 14],
    [11, 11, 11, 11, 12, 12, 13, 13, 14, 14, 15, 15],
    [12, 12, 12, 12, 12, 13, 13, 14, 14, 15, 15, 15],
]

# ─────────────────────────────────────────────────────────────────────────────
# SCORING + ANGLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def risk_cat(s):
    if s == 1:
        return "Dapat Diabaikan", "🟢", "Tidak perlu tindakan", "#27AE60"
    elif s <= 3:
        return "Rendah", "🟡", "Perubahan mungkin diperlukan", "#2ECC71"
    elif s <= 7:
        return "Sedang", "🟠", "Investigasi & perubahan segera", "#F39C12"
    elif s <= 10:
        return "Tinggi", "🔴", "Investigasi & implementasi segera", "#E67E22"
    else:
        return "Sangat Tinggi", "🚨", "Implementasi perubahan SEGERA!", "#E74C3C"


def calc_angle(a, b, c):
    a = np.array(a, float)
    b = np.array(b, float)
    c = np.array(c, float)
    ba = a - b
    bc = c - b
    cos = np.dot(ba, bc) / (np.linalg.norm(ba) * np.linalg.norm(bc) + 1e-10)
    return round(np.degrees(np.arccos(np.clip(cos, -1, 1))), 1)


def trunk_flexion(shoulder, hip):
    v = np.array([hip[0] - shoulder[0], hip[1] - shoulder[1]])
    cos = np.dot(v, [0, 1]) / (np.linalg.norm(v) + 1e-10)
    return round(np.degrees(np.arccos(np.clip(cos, -1, 1))), 1)


def neck_flexion(ear, shoulder, hip):
    nv = np.array([ear[0] - shoulder[0], ear[1] - shoulder[1]])
    tv = np.array([shoulder[0] - hip[0], shoulder[1] - hip[1]])
    cos = np.dot(nv, tv) / (np.linalg.norm(nv) * np.linalg.norm(tv) + 1e-10)
    return round(np.degrees(np.arccos(np.clip(cos, -1, 1))), 1)


def upper_arm_angle(shoulder, elbow, hip):
    av = np.array([elbow[0] - shoulder[0], elbow[1] - shoulder[1]])
    tv = np.array([hip[0] - shoulder[0], hip[1] - shoulder[1]])
    cos = np.dot(av, tv) / (np.linalg.norm(av) * np.linalg.norm(tv) + 1e-10)
    return round(np.degrees(np.arccos(np.clip(cos, -1, 1))), 1)


def score_neck(a):
    return 1 if a <= 20 else 2


def score_trunk(a):
    if a < 5:
        return 1
    elif a <= 20:
        return 2
    elif a <= 60:
        return 3
    else:
        return 4


def score_legs(knee_angle):
    b = 1
    f = 180 - knee_angle
    if 30 <= f <= 60:
        b += 1
    elif f > 60:
        b += 2
    return min(b, 4)


def score_ua(a):
    if a <= 20:
        return 1
    elif a <= 45:
        return 2
    elif a <= 90:
        return 3
    else:
        return 4


def score_la(a):
    f = 180 - a
    return 1 if 60 <= f <= 100 else 2


def score_wrist(d):
    return 1 if d <= 15 else 2


def tbl_a(t, n, l):
    return TABLE_A[max(1, min(5, t)) - 1][max(1, min(3, n)) - 1][max(1, min(4, l)) - 1]


def tbl_b(u, l, w):
    return TABLE_B[max(1, min(6, u)) - 1][max(1, min(2, l)) - 1][max(1, min(3, w)) - 1]


def tbl_c(a, b):
    return TABLE_C[max(1, min(12, a)) - 1][max(1, min(12, b)) - 1]


def force_score(kg):
    try:
        v = float(kg)
    except Exception:
        v = 0
    return 0 if v < 5 else (1 if v <= 10 else 2)


# ─────────────────────────────────────────────────────────────────────────────
# SIMPLE OVERLAY (ringan)
# ─────────────────────────────────────────────────────────────────────────────
def draw_overlay(img_bgr, pts, reba_score, category, color_hex):
    out = img_bgr.copy()
    color_hex = color_hex.lstrip("#")
    bgr = (int(color_hex[4:6], 16), int(color_hex[2:4], 16), int(color_hex[0:2], 16))

    pairs = [
        ("ear", "shoulder"),
        ("shoulder", "elbow"),
        ("elbow", "wrist"),
        ("shoulder", "hip"),
        ("hip", "knee"),
        ("knee", "ankle"),
    ]
    for a, b in pairs:
        if a in pts and b in pts:
            p1 = tuple(np.array(pts[a], int))
            p2 = tuple(np.array(pts[b], int))
            cv2.line(out, p1, p2, (0, 200, 120), 2, cv2.LINE_AA)

    for _, v in pts.items():
        p = tuple(np.array(v, int))
        cv2.circle(out, p, 4, (50, 220, 255), -1, cv2.LINE_AA)
        cv2.circle(out, p, 5, (255, 255, 255), 1, cv2.LINE_AA)

    h, w = out.shape[:2]
    cv2.rectangle(out, (0, 0), (w, 48), bgr, -1)
    cv2.putText(out, f"REBA: {reba_score} | {category}", (12, 32),
                cv2.FONT_HERSHEY_SIMPLEX, 0.85, (255, 255, 255), 2, cv2.LINE_AA)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# ANALYZE 1 IMAGE (beban & activity_score per foto)
# ─────────────────────────────────────────────────────────────────────────────
def analyze_pose(image_bgr, beban, aktivitas, activity_score):
    res = pose_detector.process(cv2.cvtColor(image_bgr, cv2.COLOR_BGR2RGB))
    if not res.pose_landmarks:
        return None, None

    h, w, _ = image_bgr.shape
    lm = res.pose_landmarks.landmark
    LP = mp_pose.PoseLandmark

    def gp(idx):
        return [lm[idx].x * w, lm[idx].y * h]

    # gunakan sisi kiri (sesuai kode kamu)
    ear = gp(LP.LEFT_EAR.value)
    nose = gp(LP.NOSE.value)
    shoulder = gp(LP.LEFT_SHOULDER.value)
    elbow = gp(LP.LEFT_ELBOW.value)
    wrist = gp(LP.LEFT_WRIST.value)
    hip = gp(LP.LEFT_HIP.value)
    knee = gp(LP.LEFT_KNEE.value)
    ankle = gp(LP.LEFT_ANKLE.value)

    ta = trunk_flexion(shoulder, hip)
    na = neck_flexion(ear, shoulder, hip)
    uaa = upper_arm_angle(shoulder, elbow, hip)
    laa = calc_angle(shoulder, elbow, wrist)
    ka = calc_angle(hip, knee, ankle)
    wr = calc_angle(elbow, wrist, shoulder)
    wd = round(abs(180 - wr), 1)

    ns = score_neck(na)
    ts = score_trunk(ta)
    ls = score_legs(ka)
    us = score_ua(uaa)
    las = score_la(laa)
    ws = score_wrist(wd)

    tA = tbl_a(ts, ns, ls)
    tB = tbl_b(us, las, ws)
    fs = force_score(beban)

    sA = tA + fs
    sB = tB + 1  # coupling default: 1 (Fair)
    sC = tbl_c(sA, sB)

    final = max(1, min(15, sC + int(activity_score)))
    kat, icon, tindakan, warna = risk_cat(final)

    pts = {
        "nose": nose,
        "ear": ear,
        "shoulder": shoulder,
        "elbow": elbow,
        "wrist": wrist,
        "hip": hip,
        "knee": knee,
        "ankle": ankle,
    }

    annotated = draw_overlay(image_bgr, pts, final, kat, warna)

    result = dict(
        Timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        Beban_kg=float(beban),
        Aktivitas=aktivitas,
        Activity_Score=int(activity_score),

        Sudut_Leher=na,
        Sudut_Tubuh=ta,
        Sudut_LenganAtas=uaa,
        Sudut_LenganBawah=laa,
        Deviasi_Pergelangan=wd,
        Sudut_Lutut=ka,

        Skor_Leher=ns,
        Skor_Tubuh=ts,
        Skor_Kaki=ls,
        Skor_LenganAtas=us,
        Skor_LenganBawah=las,
        Skor_Pergelangan=ws,

        Table_A=tA,
        Table_B=tB,
        Force_Score=fs,
        Score_A=sA,
        Score_B=sB,
        Score_C=sC,

        REBA_Final=final,
        Kategori=kat,
        Tindakan=tindakan,
        Warna=warna,
        Icon=icon,
    )

    return annotated, result


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT (batch)
# ─────────────────────────────────────────────────────────────────────────────
def build_excel_batch(results: list, nama_laporan: str) -> bytes:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(results)

    prefer = [
        "Nama File", "Timestamp", "Aktivitas", "Beban_kg", "Activity_Score",
        "REBA_Final", "Kategori", "Tindakan",
        "Sudut_Leher", "Sudut_Tubuh", "Sudut_LenganAtas", "Sudut_LenganBawah",
        "Deviasi_Pergelangan", "Sudut_Lutut",
        "Skor_Leher", "Skor_Tubuh", "Skor_Kaki", "Skor_LenganAtas",
        "Skor_LenganBawah", "Skor_Pergelangan",
        "Table_A", "Force_Score", "Score_A", "Table_B", "Score_B", "Score_C",
    ]
    cols = [c for c in prefer if c in df.columns] + [c for c in df.columns if c not in prefer]
    df = df[cols]

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as wr:
        df.to_excel(wr, sheet_name="Hasil REBA", index=False, startrow=2)
        ws = wr.sheets["Hasil REBA"]

        ws.merge_cells("A1:H1")
        ws["A1"] = f"Laporan REBA (Batch) — {nama_laporan} — {datetime.now().strftime('%d %B %Y')}"
        ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1A3A6A")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        for cell in ws[3]:
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="2E4A7A")
            cell.alignment = Alignment(horizontal="center")

        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        max_row = 3 + len(df)
        max_col = len(df.columns)
        for r in range(3, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(r, c).border = thin

        for col in ws.columns:
            ml = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(ml + 3, 12)

        if "REBA_Final" in df.columns:
            reba_idx = df.columns.get_loc("REBA_Final") + 1
            cmap = {(1, 1): "27AE60", (2, 3): "2ECC71", (4, 7): "F39C12", (8, 10): "E67E22", (11, 15): "E74C3C"}
            for rr in range(4, 4 + len(df)):
                val = ws.cell(rr, reba_idx).value
                try:
                    score = int(val)
                except Exception:
                    continue
                cc = "FFFFFF"
                for (lo, hi), hx in cmap.items():
                    if lo <= score <= hi:
                        cc = hx
                        break
                ws.cell(rr, reba_idx).fill = PatternFill("solid", fgColor=cc)
                ws.cell(rr, reba_idx).font = Font(bold=True, color="000000")

    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
for k, d in [
    ("inputs_df", None),
    ("batch_results", []),
    ("batch_images", []),
    ("batch_failed", []),
    ("analyzed", False),
    ("selected_file", None),
]:
    if k not in st.session_state:
        st.session_state[k] = d


def _make_inputs_df(files):
    rows = []
    for i, f in enumerate(files, start=1):
        file_key = f"{i:03d}_{f.name}"
        rows.append({
            "File Key": file_key,
            "Nama File": f.name,
            "Beban_kg": 0.0,
            "Activity_Score": 0,
            "Aktivitas": "",
        })
    return pd.DataFrame(rows)


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("🦴 REBA Analyzer")
    st.caption("Ergonomics Risk Assessment · Hignett & McAtamney, 2000")
    st.divider()

    st.subheader("📁 Upload Foto (Multi)")
    uploaded_files = st.file_uploader(
        "Pilih foto postur kerja (boleh banyak sekaligus)",
        type=["jpg", "jpeg", "png", "bmp"],
        accept_multiple_files=True,
        help="Pastikan seluruh tubuh (kepala hingga kaki) terlihat jelas",
    )

    st.divider()
    st.subheader("⚙️ Default untuk Semua Foto (opsional)")
    default_beban = st.number_input("Default Beban (kg)", min_value=0.0, max_value=500.0, value=0.0, step=0.5)
    default_as = st.number_input("Default Activity Score (0–3)", min_value=0, max_value=3, value=0, step=1)
    default_aktiv = st.text_input("Default Aktivitas", value="")

    apply_all = st.button("Terapkan default ke semua foto", use_container_width=True, disabled=(not uploaded_files))

    st.divider()
    analyze_btn = st.button(
        "🔍 Analisis REBA (Batch)",
        type="primary",
        use_container_width=True,
        disabled=(not uploaded_files),
    )

    st.divider()
    if st.button("🚪 Logout", use_container_width=True):
        st.session_state.logged_in = False
        st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN AREA
# ─────────────────────────────────────────────────────────────────────────────
st.title("🦴 REBA Ergonomi — Batch Upload")

col_left, col_right = st.columns([3, 2], gap="large")

# Build/refresh inputs_df when uploads change
if uploaded_files:
    current_keys = [f"{i:03d}_{f.name}" for i, f in enumerate(uploaded_files, start=1)]

    if st.session_state.inputs_df is None:
        st.session_state.inputs_df = _make_inputs_df(uploaded_files)
    else:
        # if file list changed, rebuild
        existing_keys = st.session_state.inputs_df["File Key"].tolist() if "File Key" in st.session_state.inputs_df.columns else []
        if existing_keys != current_keys:
            st.session_state.inputs_df = _make_inputs_df(uploaded_files)

    if apply_all and st.session_state.inputs_df is not None:
        df0 = st.session_state.inputs_df.copy()
        df0["Beban_kg"] = float(default_beban)
        df0["Activity_Score"] = int(default_as)
        if default_aktiv.strip():
            df0["Aktivitas"] = default_aktiv.strip()
        st.session_state.inputs_df = df0

with col_left:
    st.subheader("1) Input per Foto (Beban & Activity Score)")
    if not uploaded_files:
        st.info("⬅️ Upload beberapa foto dulu di sidebar.")
    else:
        st.caption("Edit nilai per baris. Kolom **Beban_kg** dan **Activity_Score** bisa berbeda tiap foto.")
        edited = st.data_editor(
            st.session_state.inputs_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            disabled=["File Key", "Nama File"],
            column_config={
                "Beban_kg": st.column_config.NumberColumn("Beban_kg", min_value=0.0, max_value=500.0, step=0.5),
                "Activity_Score": st.column_config.NumberColumn("Activity_Score", min_value=0, max_value=3, step=1),
                "Aktivitas": st.column_config.TextColumn("Aktivitas"),
            },
            key="inputs_editor",
        )
        # Update session state with edited df
        st.session_state.inputs_df = edited

    st.divider()
    st.subheader("3) Visualisasi Skeleton")

    if st.session_state.analyzed and st.session_state.batch_images:
        file_names = [x[0] for x in st.session_state.batch_images]
        if st.session_state.selected_file not in file_names:
            st.session_state.selected_file = file_names[0]
        pick = st.selectbox("Pilih file untuk ditampilkan", file_names, index=file_names.index(st.session_state.selected_file))
        st.session_state.selected_file = pick
        img = dict(st.session_state.batch_images).get(pick)
        if img is not None:
            st.image(cv2.cvtColor(img, cv2.COLOR_BGR2RGB), use_container_width=True, caption=f"Skeleton — {pick}")
        if st.session_state.batch_failed:
            st.warning("Pose tidak terdeteksi pada: " + ", ".join(st.session_state.batch_failed))
    else:
        if uploaded_files:
            st.image(Image.open(uploaded_files[0]).convert("RGB"), use_container_width=True,
                     caption="Pratinjau file pertama — setelah Analisis, pilih file untuk lihat overlay")

with col_right:
    st.subheader("2) Hasil & Export")

    if not (st.session_state.analyzed and st.session_state.batch_results):
        st.info("Klik **Analisis REBA (Batch)** setelah input per foto selesai.", icon="📋")
    else:
        df = pd.DataFrame(st.session_state.batch_results)
        show_cols = [c for c in ["Nama File", "Beban_kg", "Activity_Score", "REBA_Final", "Kategori", "Tindakan", "Aktivitas"] if c in df.columns]
        st.markdown("**📊 Ringkasan Batch**")
        st.dataframe(df[show_cols], use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("**📤 Export Excel (Semua Foto)**")
        nama_laporan = st.text_input("Nama laporan batch", value="REBA Batch", key="nama_laporan_batch")
        if nama_laporan.strip():
            st.download_button(
                label="⬇️ Download Excel (Batch)",
                data=build_excel_batch(st.session_state.batch_results, nama_laporan.strip()),
                file_name=f"REBA_Batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


# ─────────────────────────────────────────────────────────────────────────────
# RUN ANALYSIS (BATCH)
# ─────────────────────────────────────────────────────────────────────────────
if analyze_btn and uploaded_files:
    # Map file_key -> UploadedFile
    file_map = {f"{i:03d}_{f.name}": f for i, f in enumerate(uploaded_files, start=1)}

    inputs_df = st.session_state.inputs_df.copy() if st.session_state.inputs_df is not None else _make_inputs_df(uploaded_files)

    batch_results = []
    batch_images = []
    batch_failed = []

    with st.spinner("Mendeteksi pose & menghitung REBA (batch)..."):
        for _, row in inputs_df.iterrows():
            file_key = row["File Key"]
            fname = row["Nama File"]
            beban = float(row.get("Beban_kg", 0.0) or 0.0)
            activity_score = int(row.get("Activity_Score", 0) or 0)
            aktivitas = str(row.get("Aktivitas", "") or "").strip()

            f = file_map.get(file_key)
            if f is None:
                continue

            img_bgr = cv2.cvtColor(np.array(Image.open(f).convert("RGB")), cv2.COLOR_RGB2BGR)
            annotated, result = analyze_pose(img_bgr, beban, aktivitas, activity_score)

            if result is None:
                batch_failed.append(fname)
                continue

            result["Nama File"] = fname
            batch_results.append(result)
            batch_images.append((fname, annotated))

    if not batch_results:
        st.error(
            "**Tidak ada pose yang berhasil terdeteksi dari semua foto.**\n\n"
            "Pastikan seluruh tubuh (kepala–kaki) terlihat jelas, pencahayaan cukup, dan tidak terhalang.",
            icon="⚠️",
        )
        st.session_state.analyzed = False
        st.session_state.batch_results = []
        st.session_state.batch_images = []
        st.session_state.batch_failed = batch_failed
    else:
        st.session_state.batch_results = batch_results
        st.session_state.batch_images = batch_images
        st.session_state.batch_failed = batch_failed
        st.session_state.analyzed = True
        st.session_state.selected_file = batch_images[0][0]
        st.rerun()
